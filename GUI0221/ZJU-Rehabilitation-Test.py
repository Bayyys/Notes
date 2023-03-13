import tkinter as tk
from tkinter import *
import serial
import serial.tools.list_ports
from tkinter import ttk
import serialPortFile
import time
import dataprocess
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.image as mpimg
import matplotlib.transforms as tr
import threading
import twoplot_new
import getpicture
import openpyxl
from collections import deque
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk

serial_list = []
stinum = {
    "0": '0'.encode("utf-8"),    # or chr(0x30).encode("utf-8")
    "1": '1'.encode("utf-8"),
    "2": '2'.encode("utf-8"),
    "3": '3'.encode("utf-8"),
    "4": '4'.encode("utf-8"),
    "5": '5'.encode("utf-8"),
    "6": '6'.encode("utf-8"),
    "7": '7'.encode("utf-8"),
    "8": '8'.encode("utf-8"),
    "9": '9'.encode("utf-8"),
    "0d": chr(0x0d).encode("utf-8")
}

def show():
    global ser, emgFlag
    text1.delete(1.0, END)  # 删除原来信息重新覆盖
    try:
        emgFlag = True
        serialPort_emg = combo_com.get()  # 串口    每触发一次， 开启一次串口再立马关闭
        baudRate_emg = combo_bps.get()
        ss = serial.Serial(serialPort_emg, baudRate_emg)
        serial_list.append(ss)
        ser = serial_list[-1]
        time.sleep(0.1)
    except:
        text1.delete('1.0', 'end')
        text1.insert(INSERT, "\n串口被占用或串口参数设置不合理")
        ser.flushInput()  # 清空串口缓存
    else:
        test_healthyside()
def show_plot():
    global ser
    try:
        # 打开串口，并得到串口对象
        serialPort_plot = combo_com.get()  # 串口    每触发一次， 开启一次串口再立马关闭
        baudRate_plot = combo_bps.get()
        ss = serial.Serial(serialPort_plot, baudRate_plot)
        serial_list.append(ss)
        ser = serial_list[-1]
        # ser = serial.Serial(serialPort, baudRate)
        # 避免上次关闭窗口可能出现的未传送完整数据，而板子还在等待od的出现
        time.sleep(0.1)
    except:
        text3.delete('1.0', 'end')
        text3.insert(INSERT, "\n串口被占用或串口参数设置不合理")
        ser.flushInput()  # 清空串口缓存
        # print(traceback.format_exc())  # 查看报错信息
    else:
        # imu_on.config(text="退出动态展示", command=close_plot)
        tk.Button(frame7, text="退出动态展示", font=(10), width=13, command=close_plot).grid(row=21, column=4, padx=5, pady=5)
        test_twoplot()


def plot_durations3(y1,y2,y3):
    y = []
    y.append(np.array([y1,y2,y3]))
    plt.figure(3)
    plt.clf()
    plt.subplot(311)
    plt.ylabel("EMG/uV")
    plt.plot(y1)
    plt.subplot(312)
    plt.ylabel("filtered EMG/uV")
    plt.plot(y2)
    plt.subplot(313)
    plt.ylabel("RMS/uV")
    plt.plot(y3)
    plt.pause(0.01)
def close():
    global emgFlag
    try:
        ser_now = serial_list[-1]
        ser_now.flushOutput()  # 终止当前写操作，并丢弃发送缓存中的数据。
        text1.delete('1.0', 'end')
        text1.insert(INSERT, "\n已退出肌电检测！")
        emgFlag = False
        plt.close()
        ser.close()
    except:
        text1.delete('1.0', 'end')
        text1.insert(INSERT, "\n已退出肌电检测（可以再次点击\"开始检测\"按钮！）；")
def close_plot():
    try:
        ser_now = serial_list[-1]
        ser_now.flushOutput()  # 终止当前写操作，并丢弃发送缓存中的数据。
        text3.delete('1.0', 'end')
        text3.insert(INSERT, "\n已退出动态展示！")
        plt.close()
        ser.close()
    except:
        text3.delete('1.0', 'end')
        text3.insert(INSERT, "\n已退出动态展示（可以再次点击\"开启动态展示\"按钮！）；")
        # imu_on.config(text="关节动态展示", command=show_plot)

def test_healthyside():  # 1.进行健侧肌电电刺激反馈阶段
    global Qmax, Qmin, Imax
    global ser, data_health_rms, emgFlag
    ser.write([0XFE, 0X04, 0X00, 0X01, 0X00, 0X04, 0X08, 0XCB, 0XD6, 0X16])
    data_health_usefully, data_health_usefully_rms, data_health_usefully_afterfilter = [], [], []
    time.sleep(3)
    ser.flushInput()
    while emgFlag:
        time_start = time.time()
        count = ser.inWaiting()  # 获取串口缓冲区数据
        if count != 0:
            recv = ser.read(ser.in_waiting).hex()
            data_health = dataprocess.decode(recv)
            print(f"1:::{recv}")
            if data_health == None or data_health[3] == []:
                data_health_usefully_rms.append(0)
            else:
                data_health_usefully_rms.append(data_health[3])
                data_health_rms = max(data_health_usefully_rms)
                data_health_usefully.extend(data_health[2])
                data_health_usefully_afterfilter.extend(data_health[4])
            plt.ion()
            plot_durations3(data_health_usefully, data_health_usefully_afterfilter, data_health_usefully_rms)
        time.sleep(0.1)  # 延时0.1秒，免得CPU出问题
        time_end = time.time()

# 进行imu检测
def test_twoplot():
    # 引入全局变量
    global ser
    global data_everychannel, data_record
    global x_coxa, y_coxa, x_knee0, y_knee0
    global x_leg1place, y_leg1place, x_leg2place0, y_leg2place0, x_dotplace0, y_dotplace0
    global EdgeFlag, EdgeFlag2, length
    # 读取腿部图片
    leg1 = getpicture.readleg1()
    leg2 = getpicture.readleg2()
    dot = getpicture.readdot()
    bkg = getpicture.readbkg()
    # 创建画图fig
    fig = plt.figure(figsize=(6, 6))
    ax = fig.add_subplot()
    # 开启交互模式
    plt.ion()
    ser.write([0XFE,0X09, 0X00, 0X01, 0X00, 0X04, 0X08, 0XCB, 0XD6, 0X16])
    time.sleep(3)
    ser.flushInput() #  清空串口缓存
    while True:
        count = ser.inWaiting()  # 获取串口缓冲区数据
        if count != 0:
            recv = ser.read(ser.in_waiting).hex()
            print(f"recv = {recv}")
            data_everychannel = twoplot_new.decode(recv)
            data_record.append(data_everychannel)
            # 清除上一帧画面并重建
            plt.clf()
            ax.figure.figimage(bkg)
            # 读取大腿\小腿姿态角
            roll = 40 - data_everychannel[8]  # roll
            roll2 = - data_everychannel[17]  # roll2
            # # 超量程换算
            # if (roll > 180 or roll < -180 or roll2 > 180 or roll2 < -180):
            #     continue
            # if data_record[-2][6] < -60 and data_record[-2][6] > -90:
            #     EdgeFlag = True
            # if data_record[-2][6] < 20 and data_record[-2][6] > -20 and data_record[-2][6] == data_record[-3][6]:
            #     EdgeFlag = False
            # if EdgeFlag and roll > 0:
            #     roll = roll - 180
            # if data_record[-2][15] < -60 and data_record[-2][15] > -90:
            #     EdgeFlag2 = True
            # if data_record[-2][15] < 20 and data_record[-2][15] > -20 and data_record[-2][15] == data_record[-3][15]:
            #     EdgeFlag2 = False
            # if EdgeFlag2 and roll2 > 0:
            #     roll2 = roll2 - 180
            # 角度值显示
            print(f"roll = {roll}, roll2 = {roll2}")
            # 单位换算
            roll = np.deg2rad(roll)
            roll2 = np.deg2rad(roll2)
            # 关节坐标
            x1 = length * np.sin(roll)
            y1 = -length * np.cos(roll)
            x2 = x1 + length * np.sin(roll2)
            y2 = y1 - length * np.cos(roll2)
            # 大腿图片
            thigh = ax.figure.figimage(leg1, x_leg1place, y_leg1place)
            rot1 = tr.Affine2D().rotate_around(x_coxa, y_coxa, roll)
            thigh.set_transform(thigh.get_transform() + rot1)
            # 小腿图片
            if x1 < 0 and y1 < 0:
                x_leg2place = x_leg2place0 + x1 * 29
                y_leg2place = y_leg2place0 + (y1 + 4) * 32
                x_knee = x_knee0 + x1 * 29
                y_knee = y_knee0 + (y1 + 4) * 32
            elif x1 > 0 and y1 < 0:
                x_leg2place = x_leg2place0 + x1 * 31
                y_leg2place = y_leg2place0 + (y1 + 4) * 27
                x_knee = x_knee0 + x1 * 31
                y_knee = y_knee0 + (y1 + 4) * 27
            elif x1 < 0 and y1 > 0:
                x_leg2place = x_leg2place0 + x1 * 28
                y_leg2place = y_leg2place0 + (y1 + 4) * 31
                x_knee = x_knee0 + x1 * 28
                y_knee = y_knee0 + (y1 + 4) * 31
            else:
                x_leg2place = x_leg2place0 + x1 * 33
                y_leg2place = y_leg2place0 + (y1 + 4) * 30
                x_knee = x_knee0 + x1 * 33
                y_knee = y_knee0 + (y1 + 4) * 30
            shank = ax.figure.figimage(leg2, x_leg2place, y_leg2place)
            rot2 = tr.Affine2D().rotate_around(x_knee, y_knee, roll2)
            shank.set_transform(shank.get_transform() + rot2)
            # 膝盖圆点
            if x1 < 0 and y1 < 0:
                x_dotplace = x_dotplace0 + x1 * 29
                y_dotplace = y_dotplace0 + (y1 + 4) * 32
            elif x1 > 0 and y1 < 0:
                x_dotplace = x_dotplace0 + x1 * 31
                y_dotplace = y_dotplace0 + (y1 + 4) * 27
            elif x1 < 0 and y1 > 0:
                x_dotplace = x_dotplace0 + x1 * 28
                y_dotplace = y_dotplace0 + (y1 + 4) * 31
            else:
                x_dotplace = x_dotplace0 + x1 * 33
                y_dotplace = y_dotplace0 + (y1 + 4) * 30
            knee_dot = ax.figure.figimage(dot, x_dotplace, y_dotplace)
            # 绘图
            # thisx = [0, x1, x2]
            # thisy = [0, y1, y2]
            # line.set_data(thisx, thisy)
            plt.pause(0.0000001)  # 暂停一秒
            plt.ioff()
        time.sleep(0.1)  # 延时0.1秒，免得CPU出问题
# imu历史数据导出
def data_export():
    wb = openpyxl.Workbook()
    ws = wb.create_sheet('imudata')
    ws.cell(row=1, column=1).value = 'accx'
    ws.cell(row=1, column=2).value = 'accy'
    ws.cell(row=1, column=3).value = 'accz'
    ws.cell(row=1, column=4).value = 'gyrox'
    ws.cell(row=1, column=5).value = 'gyroy'
    ws.cell(row=1, column=6).value = 'gyroz'
    ws.cell(row=1, column=7).value = 'roll'
    ws.cell(row=1, column=8).value = 'pitch'
    ws.cell(row=1, column=9).value = 'yaw'
    ws.cell(row=1, column=10).value = 'accx2'
    ws.cell(row=1, column=11).value = 'accy2'
    ws.cell(row=1, column=12).value = 'accz2'
    ws.cell(row=1, column=13).value = 'gyrox2'
    ws.cell(row=1, column=14).value = 'gyroy2'
    ws.cell(row=1, column=15).value = 'gyroz2'
    ws.cell(row=1, column=16).value = 'roll2'
    ws.cell(row=1, column=17).value = 'pitch2'
    ws.cell(row=1, column=18).value = 'yaw2'
    for i in range(len(data_record)):
        for j in range(18):
            ws.cell(row=i+2, column=j+1).value = data_record[i][j]
    wb.save('D:\imudata.xlsx')
    text3.delete('1.0', 'end')
    text3.insert(INSERT, "\n保存成功！保存路径为：'D:\imudata.xlsx'")
def renew_com():
    try:
        serial_com = serialPortFile.GetCom()
        combo_com['values'] = serial_com
        combo_com.current(0)  # 目前显示的是第几个串口号
    except:
        text0.delete('1.0', 'end')
        text0.insert(INSERT, "读取串口失败，请检查串口连接")
    else:
        text0.delete('1.0', 'end')
        text0.insert(INSERT, "读取串口成功！")
def thread_it(func, *args):
    """将函数打包进线程"""
    # 创建
    t = threading.Thread(target=func, args=args)
    # 守护 !!!
    t.daemon = True
    # 启动
    t.start()
    # 阻塞--卡死界面！
    # t.join()
    
if __name__ == "__main__":
    global e11,e21,e31,e32,e33,e34,e41,e51,e61,e71,e81,e131,e141,Ir_fes
    root = tk.Tk()
    root.title("ZJU-Rehabilitation-Test")
    root.geometry('700x350+340+240')  # 规定窗口大小以及位置
    # root.resizable(False, False)  # 规定窗口不可缩放
    wid_in = 15  # 和前面一列之间的宽度
    wid_in2 = 10  # 和前面一列之间的宽度
    # -------------------------------------------------------------------------

    # -------------------------------------------------------------------------
    # 容器设置（三大板块：串口设置、肌电检测、姿态检测）
    frame_root = Frame(root)
    frame_left = Frame(frame_root)  # 左半边
    frame_right = Frame(frame_root)  # 右半边
    widd = 200
    pw1 = PanedWindow(frame_left, orient=VERTICAL, width=widd)
    frame1 = LabelFrame(pw1, text="串口设置", width=widd, padx=5, pady=5)
    frame2 = Frame(frame_left, width=widd, padx=5, pady=5)
    pw1.add(frame1, width=widd, padx=5, pady=5)
    pw1.pack(side=TOP)
    frame2.pack(side=TOP)

    pw3 = PanedWindow(frame_right, orient=VERTICAL, width=widd)
    frame5 = LabelFrame(pw3, text="肌电检测", width=widd, padx=10, pady=5)
    frame6 = Frame(frame_right, width=widd, padx=10, pady=5)
    pw3.add(frame5, width=widd)
    pw3.pack(side=TOP)
    frame5.pack(side=TOP)

    pw4 = PanedWindow(frame_right, orient=VERTICAL, width=widd)
    frame7 = LabelFrame(pw4, text="姿态检测", width=widd, padx=10, pady=5)
    frame8 = Frame(frame_right, width=widd, padx=10, pady=5)
    pw4.add(frame7, width=widd)
    pw4.pack(side=TOP)
    frame7.pack(side=TOP)
    canvas = tk.Canvas() # 创建画布控件canvas
    # -------------------------------------------------------------------------
    # 按键与提示框
    tk.Button(frame1, text="刷新", font=(5), width=8, command=lambda: thread_it(renew_com)).grid(row=5, column=0, padx=2,pady=2)
    text0 = tk.Text(frame1, width=21, height=3)
    text0.grid(row=6, column=0, padx=10, pady=5, columnspan=3)
    # 串口设置区
    label_com = Label(frame1, text="串口号", height=2, width=wid_in2).grid(row=0, column=0, padx=1, pady=1)
    label_bps = Label(frame1, text="波特率", height=2).grid(column=0, row=1)
    label_datBit = Label(frame1, text="数据位", height=2).grid(column=0, row=2)
    label_parity = Label(frame1, text="校验位", height=2).grid(column=0, row=3)
    label_stop_bit = Label(frame1, text="停止位", height=2).grid(column=0, row=4)

    # 串口号
    # 接下来是串口号的combobox的设置，其中serialPortFile.GetCom()是获取所有串口号的函数，接下来会讲解
    varPort = StringVar()
    combo_com = ttk.Combobox(frame1, textvariable=varPort, width=8, height=2, justify=CENTER)  # 下拉下选框的宽度和高度
    serial_com = serialPortFile.GetCom()
    combo_com['values'] = serial_com
    # combo_com.bind("<<ComboboxSelected>>", lambda event: combo1_handler(var=varPort.get()))
    try:
        combo_com.current(0)  # 目前显示的是第几个串口号
    except:
        text0.delete('1.0', 'end')
        text0.insert(INSERT, "请插入USB并点击“刷新”读取串口")
    else:
        text0.delete('1.0', 'end')
        text0.insert(INSERT, "下拉选框可进行串口设置")
    combo_com.grid(column=1, row=0)

    # 波特率
    varBitrate = StringVar()
    combo_bps = ttk.Combobox(frame1, textvariable=varBitrate, width=8, height=2, justify=CENTER)
    combo_bps['values'] = ("115200","4608000")
    # combo_bps.bind("<<ComboboxSelected>>", lambda event: combo2_handler(var=varBitrate.get()))
    combo_bps.current(0)
    combo_bps.grid(column=1, row=1)

    # 数据位
    combo_byteBit = ttk.Combobox(frame1, width=8, height=2, justify=CENTER)
    combo_byteBit['values'] = ("5", "6", "7", "8")
    combo_byteBit.current(3)
    combo_byteBit.grid(column=1, row=2)

    # 奇偶校验
    combo_parity = ttk.Combobox(frame1, width=8, height=2, justify=CENTER)
    combo_parity['values'] = ("None", "Odd", "Even")
    combo_parity.current(0)
    combo_parity.grid(column=1, row=3)

    # 停止位
    combo_stopBit = ttk.Combobox(frame1, width=8, height=2, justify=CENTER)
    combo_stopBit['values'] = ("1", "2")
    combo_stopBit.current(0)
    combo_stopBit.grid(column=1, row=4)

    # -------------------------------------------------------------------------
    # 肌电检测设置区
    # 全局变量
    emgFlag = False
    # 按键与文本框设置
    EMGtest = tk.Button(frame5, text="开始检测", font=(10), width=13, command=lambda: thread_it(show)).grid(row=19, column=2, padx=5, pady=5)# , font=("宋体", 15)
    tk.Button(frame5, text="结束检测", font=(10), width=13, command=lambda: thread_it(close)).grid(row=19, column=3, padx=5, pady=5)
    text1 = tk.Text(frame5, width=50, height=5)
    text1.grid(row=20, column=2, padx=10, pady=5, columnspan=3)
    text1.insert(INSERT, "\n点击“开始检测”进行肌电检测")
    # -------------------------------------------------------------------------
    # IMU检测设置区
    # 全局变量
    data_everychannel = [i for i in range(18)]
    data_record = deque([[0 for i in range(18)] for j in range(3)], maxlen=1000)
    x_coxa, y_coxa, x_knee0, y_knee0 = 310, 300, 308, 182
    x_leg1place, y_leg1place, x_leg2place0, y_leg2place0, x_dotplace0, y_dotplace0 = 70, 160, 89, 42, 284, 158
    EdgeFlag, EdgeFlag2, length = False, False, 4
    # 按键设置
    imu_on = tk.Button(frame7, text="关节动态展示", font=(10), width=16, command=show_plot)
    imu_on.grid(row=21, column=2, padx=5, pady=5)
    imu_cal = tk.Button(frame7, text="导出最近检测记录", font=(10), width=16, command=lambda: thread_it(data_export))
    imu_cal.grid(row=21, column=3, padx=5, pady=5)
    # 文本框设置
    text3 = tk.Text(frame7, width=50, height=5)
    text3.grid(row=22, column=2, padx=10, pady=5, columnspan=3)
    text3.insert(INSERT, "\n点击“关节动态展示”开启腿部关节姿态展示")
    # -------------------------------------------------------------------------

    frame_left.pack(side=LEFT)
    frame_right.pack(side=RIGHT)
    frame_root.pack()
    root.mainloop()